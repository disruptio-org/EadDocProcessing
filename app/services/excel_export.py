"""Excel export — generate index.xlsx with full document indexation."""

from __future__ import annotations

from pathlib import Path

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.schemas.common import DocumentRecord
from app.storage import local as storage

logger = structlog.get_logger(__name__)

# Column definitions: (header_name, field_name, width)
COLUMNS = [
    ("source_file_id", "source_file_id", 38),
    ("doc_id", "doc_id", 38),
    ("page_start", "page_start", 12),
    ("page_end", "page_end", 12),
    ("supplier_A", "supplier_a", 25),
    ("po_primary_A", "po_primary_a", 18),
    ("po_secondary_A", "po_secondary_a", 18),
    ("confidence_A", "confidence_a", 14),
    ("method_A", "method_a", 12),
    ("supplier_B", "supplier_b", 25),
    ("po_primary_B", "po_primary_b", 18),
    ("po_secondary_B", "po_secondary_b", 18),
    ("confidence_B", "confidence_b", 14),
    ("method_B", "method_b", 12),
    ("match_status", "match_status", 16),
    ("decided_po_primary", "decided_po_primary", 20),
    ("decided_po_secondary", "decided_po_secondary", 20),
    ("status", "status", 10),
    ("next_action", "next_action", 18),
    ("reject_reason", "reject_reason", 40),
]


def generate_index_excel(
    source_file_id: str,
    documents: list[DocumentRecord],
) -> Path:
    """Generate the index.xlsx file with full document indexation.

    Args:
        source_file_id: UUID of the source file.
        documents: List of DocumentRecord with all pipeline results.

    Returns:
        Path to the generated Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Document Index"

    # --- Header row ---
    header_font = Font(bold=True, size=11)
    for col_idx, (header, _field, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Data rows ---
    for row_idx, doc in enumerate(documents, start=2):
        doc_dict = doc.model_dump() if hasattr(doc, "model_dump") else doc.__dict__
        for col_idx, (_header, field, _width) in enumerate(COLUMNS, start=1):
            value = doc_dict.get(field)
            # Convert enums to their string value
            if hasattr(value, "value"):
                value = value.value
            ws.cell(row=row_idx, column=col_idx, value=value)

    # --- Freeze first row ---
    ws.freeze_panes = "A2"

    # --- Save ---
    excel_path = storage.get_excel_path(source_file_id)
    wb.save(str(excel_path))

    logger.info(
        "excel_generated",
        source_file_id=source_file_id,
        rows=len(documents),
        path=str(excel_path),
    )
    return excel_path
